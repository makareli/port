# moduls and stuff
import datetime

from openpyxl import load_workbook
from countries import COUNTRY_LIST
from countryinfo import CountryInfo

print("welcome to Shahar's port!")


# menu of the program
def menu():
    try:
        print("\nLA MENU:\ntype any number for requested result :)")
        choice = int(input("for registaring a new ship type: 1.\n"
                           "to leave the docks type: 2.\n"
                           "for information regarding the ship/s type: 3.\n"
                           "to exit the program type 4.\n"
                           "your answer: "))
        if choice == 1:
            Ship()
            leave_program()
        elif choice == 2:
            search_by = int(input("please enter ship number: "))
            wb = load_workbook("Namal.xlsx")
            ws = wb.active
            flag = 0
            for row in ws.iter_rows():
                for cell in row:
                    if ws.cell(row=cell.row, column=1).value == search_by:
                        flag += 1
                        if ws.cell(row=cell.row, column=7).value != "Left":
                            ws.cell(row=cell.row, column=7).value = "Left"
                            ws.cell(row=cell.row, column=6).value = datetime.datetime.now()
                            wb.save("Namal.xlsx")
                            print("activity has been updated!")
                            break
                        else:
                            print("the ship has already left the port!")
                            break
            if flag == 0:
                print("ship number doesn't exist.")
            leave_program()
        elif choice == 3:
            search_by = input("identify ship by any detail. your answer: ").capitalize()
            wb = load_workbook("Namal.xlsx")
            ws = wb.active
            flag = 0
            nation = None
            for row in ws.iter_rows():
                for cell in row:
                    if str(cell.value) == search_by:
                        for i in range(1, 8):
                            if i == 3:
                                nation = CountryInfo(ws.cell(row=cell.row, column=i).value)
                                print(ws.cell(row=cell.row, column=i).value, end=" ")
                            elif ws.cell(row=cell.row, column=i).value == "Arrived" or\
                                    ws.cell(row=cell.row, column=i).value == "Left":
                                print(ws.cell(row=cell.row, column=i).value, end=" ")
                                print("")
                            else:
                                print(ws.cell(row=cell.row, column=i).value, end=" ")
                        flag += 1
            if flag >= 1:
                print("")
                yes_or_no = input("would you like to get information regarding the ship's country (yes / no)?"
                                  " your answer: ")
                if yes_or_no.lower() == 'yes':
                    print(nation.info())
                else:
                    print("ok.")
                leave_program()
            else:
                print("ship doesn't exist")
                print("")
                leave_program()
        elif choice == 4:
            print("\nthanks for using Shahar's port! goodbye :)")
            exit()
        else:
            print("\nnumber's not good :) please try again.\n")
            menu()
    except ValueError as e:
        print(e, "\n\nplease try again.\n")
        menu()


# creating an object of a ship and adding it to the Excel file, stupid thing but too late to change
class Ship:
    def __init__(self):
        try:
            while True:
                ship_num = int(input("enter ship's number (has to be six digits): "))
                if ship_num_exists(ship_num):
                    if 000000 <= ship_num <= 999999:
                        self.ship_num = ship_num
                        break
                    print("ship number isn't six digits. please try again.")
                else:
                    print("ship number already exists. please try again.")
            self.ship_name = input("enter ship's name: ")
            while True:
                nation = input("enter ship's country: ")
                if nation.capitalize() in COUNTRY_LIST:
                    self.ship_flag = CountryInfo(nation)
                    break
                print("no such country, try again!")
            self.date = datetime.datetime.now()
            self.activity = "arrived"
            cargo = input("enter ship's cargo: ")
            if len(cargo) > 6 and cargo[-1] == 'x'.lower() and self.ship_flag.name().lower() == 'iran':
                while True:
                    yes_or_no = input("dangerous ship! are you sure you want to accept it? (yes / no): ")
                    if yes_or_no == 'yes'.lower():
                        print("ok ship will be added.")
                        break
                    elif yes_or_no == 'no'.lower():
                        print("ship won't be added")
                        leave_program()
                        break
                    else:
                        print("didn't answer my question you twat.")
            self.cargo = cargo.capitalize()
            wb = load_workbook("Namal.xlsx")
            ws = wb.active
            new_row_data = [self.ship_num, self.ship_name.capitalize(), self.ship_flag.name().capitalize(), self.cargo,
                            self.date, "/", self.activity.capitalize()]
            ws.append(new_row_data)
            wb.save("Namal.xlsx")
            print("great! welcome to Shahar's port :)")
        except ValueError as e:
            print(e, "\nfuck you troll!")


# a function that determines if a ship number already exists
def ship_num_exists(ship_num):
    wb = load_workbook("Namal.xlsx")
    ws = wb.active
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            if ws.cell(row=cell.row, column=1).value == ship_num:
                count += 1
            else:
                pass
    if count > 0:
        return False
    else:
        return True


# function that offers you to exit the program
def leave_program():
    yes_or_no = input("would you like to leave the program (yes/no): ")
    if yes_or_no.lower() == 'no':
        print("\n\n\n")
        menu()
    elif yes_or_no.lower() == 'yes':
        print("\nthanks for using Shahar's port! goodbye :)")
        exit()
    else:
        print("not the answer I wanted, please answer the question you twat :(")
        leave_program()


menu()
